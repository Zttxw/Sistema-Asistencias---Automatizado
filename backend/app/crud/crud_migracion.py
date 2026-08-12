import io
import csv
from datetime import datetime, date, time as time_type
from typing import Dict, Any, List, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.empleado import Empleado
from app.models.asistencia import Asistencia


def generar_plantilla_excel() -> bytes:
    """
    Genera un archivo Excel in-memory (.xlsx) formateado como plantilla para la migración.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Migración"

    # Estilos
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "DNI_Practicante",
        "Fecha (YYYY-MM-DD)",
        "Hora_Entrada (HH:MM)",
        "Hora_Salida (HH:MM)"
    ]

    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filas de ejemplo (Lunes a Viernes)
    ejemplos = [
        ["72839401", "2026-08-03", "08:00", "14:00"],
        ["72839401", "2026-08-04", "08:15", "14:30"],
        ["84920192", "2026-08-05", "08:00", "13:00"],
    ]

    for row_data in ejemplos:
        ws.append(row_data)

    # Formatear filas de datos
    for row in ws.iter_rows(min_row=2, max_row=len(ejemplos)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 22)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generar_plantilla_csv() -> bytes:
    """
    Genera un archivo CSV in-memory (.csv) formateado con UTF-8 BOM para fácil apertura en Excel.
    """
    output = io.StringIO()
    output.write('\ufeff')  # UTF-8 BOM
    writer = csv.writer(output, delimiter=';')
    writer.writerow(["DNI_Practicante", "Fecha (YYYY-MM-DD)", "Hora_Entrada (HH:MM)", "Hora_Salida (HH:MM)"])
    writer.writerow(["72839401", "2026-08-03", "08:00", "14:00"])
    writer.writerow(["72839401", "2026-08-04", "08:15", "14:30"])
    writer.writerow(["84920192", "2026-08-05", "08:00", "13:00"])
    return output.getvalue().encode('utf-8')


def _normalize_dni(val: Any) -> str:
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    elif "." in val_str:
        val_str = val_str.split(".")[0]
    return val_str.strip()


from app.utils.holidays import obtener_nombre_feriado, FERIADOS_FIJOS, FERIADOS_VARIABLES


def purgar_asistencias_migradas(
    db: Session,
    empleado_id: Optional[int] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None
) -> Dict[str, Any]:
    """
    Elimina únicamente los registros de asistencia cuyo origen sea 'migracion' o agente 'MIGRACION_ADMIN'.
    Protege los registros generados por el agente de red ARP u otros métodos ('automatico', 'manual').
    Verifica además que el período no haya sido firmado previamente por Jefatura.
    """
    from app.models.informe_firmado import InformeFirmado

    query = db.query(Asistencia).filter(
        (Asistencia.origen_entrada == "migracion") | (Asistencia.agente_id == "MIGRACION_ADMIN")
    )

    if empleado_id:
        query = query.filter(Asistencia.empleado_id == empleado_id)

    if fecha_inicio:
        query = query.filter(Asistencia.fecha >= fecha_inicio)

    if fecha_fin:
        query = query.filter(Asistencia.fecha <= fecha_fin)

    asistencias_a_borrar = query.all()
    if not asistencias_a_borrar:
        return {
            "ok": True,
            "eliminados": 0,
            "message": "No se encontraron registros de migración para eliminar en el criterio seleccionado."
        }

    # Verificar si alguna asistencia pertenece a un período firmado por la Jefatura
    for asis in asistencias_a_borrar:
        informe = db.query(InformeFirmado).filter(
            InformeFirmado.empleado_id == asis.empleado_id,
            InformeFirmado.semana_inicio <= asis.fecha,
            InformeFirmado.semana_fin >= asis.fecha,
            InformeFirmado.firmado == True
        ).first()
        if informe:
            raise ValueError(
                f"No se puede purgar la asistencia del día {asis.fecha} porque pertenece a un informe mensual ya firmado."
            )

    count = len(asistencias_a_borrar)
    for asis in asistencias_a_borrar:
        db.delete(asis)

    db.commit()
    return {
        "ok": True,
        "eliminados": count,
        "message": f"Se eliminaron exitosamente {count} registros de migración."
    }


def _parse_date(val: Any) -> date:
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    val_str = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Formato de fecha inválido '{val_str}'. Use YYYY-MM-DD.")


def _parse_time(val: Any) -> time_type:
    if isinstance(val, time_type):
        return val
    if isinstance(val, datetime):
        return val.time()
    val_str = str(val).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(val_str, fmt).time()
        except ValueError:
            pass
    raise ValueError(f"Formato de hora inválido '{val_str}'. Use HH:MM (ej. 08:30).")


def _extraer_filas_desde_csv(file_bytes: bytes) -> List[Tuple[int, List[Any]]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("No se pudo decodificar el archivo CSV. Verifique la codificación.")

    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []

    sample = "\n".join(lines[:5])
    delimiter = ';'
    if ';' in sample:
        delimiter = ';'
    elif ',' in sample:
        delimiter = ','
    elif '\t' in sample:
        delimiter = '\t'

    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample, delimiters=';,\t|')
        delimiter = dialect.delimiter
    except Exception:
        pass

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    filas = []
    for num_fila, row in enumerate(reader, start=1):
        if num_fila == 1:
            # omitir encabezado
            continue
        if row and any(cell.strip() for cell in row):
            filas.append((num_fila, row))
    return filas


def _extraer_filas_desde_excel(file_bytes: bytes) -> List[Tuple[int, List[Any]]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    filas = []
    for num_fila, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and any(row):
            filas.append((num_fila, list(row)))
    return filas


def procesar_migracion_archivo(db: Session, file_bytes: bytes, filename: str, fecha_limite: date) -> Dict[str, Any]:
    """
    Procesa un archivo Excel (.xlsx / .xls) o CSV (.csv) de migración masiva e inserta/actualiza las asistencias históricas.
    """
    fn_lower = filename.lower()
    filas_con_num: List[Tuple[int, List[Any]]] = []

    if fn_lower.endswith('.csv'):
        try:
            filas_con_num = _extraer_filas_desde_csv(file_bytes)
        except Exception as e:
            return {"ok": False, "error": f"Error al leer archivo CSV: {str(e)}"}
    else:
        try:
            filas_con_num = _extraer_filas_desde_excel(file_bytes)
        except Exception as ex_excel:
            # Fallback a intentar CSV por si renombraron el archivo
            try:
                filas_con_num = _extraer_filas_desde_csv(file_bytes)
            except Exception:
                return {"ok": False, "error": f"Error al abrir el archivo Excel: {str(ex_excel)}"}

    filas_procesadas = 0
    filas_creadas = 0
    filas_actualizadas = 0
    filas_omitidas_fecha = 0
    errores: List[Dict[str, Any]] = []

    # Cache de empleados por DNI normalizado
    empleados_db = db.query(Empleado).all()
    mapa_empleados = {_normalize_dni(emp.documento): emp for emp in empleados_db if emp.documento}

    for num_fila, row in filas_con_num:
        dni_val = row[0] if len(row) > 0 else None
        fecha_val = row[1] if len(row) > 1 else None
        entrada_val = row[2] if len(row) > 2 else None
        salida_val = row[3] if len(row) > 3 else None

        if not dni_val or not fecha_val or not entrada_val:
            errores.append({
                "fila": num_fila,
                "error": "Campos obligatorios incompletos (DNI, Fecha o Hora Entrada)."
            })
            continue

        dni_str = _normalize_dni(dni_val)
        empleado = mapa_empleados.get(dni_str)
        if not empleado:
            errores.append({
                "fila": num_fila,
                "dni": dni_str,
                "error": f"No existe ningún practicante registrado con el DNI '{dni_str}'."
            })
            continue

        try:
            fecha_rec = _parse_date(fecha_val)
        except ValueError as ve:
            errores.append({"fila": num_fila, "dni": dni_str, "error": str(ve)})
            continue

        # Validación de días laborables (Solo Lunes a Viernes: 0 a 4)
        DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        if fecha_rec.weekday() >= 5:  # Sábado (5) o Domingo (6)
            filas_omitidas_fecha += 1
            nombre_dia = DIAS_SEMANA[fecha_rec.weekday()]
            errores.append({
                "fila": num_fila,
                "dni": dni_str,
                "error": f"La fecha {fecha_rec} es {nombre_dia} (fin de semana). Solo se migran asistencias de Lunes a Viernes."
            })
            continue

        # Validación de Feriados Calendario Nacionales
        nombre_feriado = obtener_nombre_feriado(fecha_rec)
        if nombre_feriado:
            filas_omitidas_fecha += 1
            errores.append({
                "fila": num_fila,
                "dni": dni_str,
                "error": f"La fecha {fecha_rec} es feriado calendario ({nombre_feriado}). Omitida de la migración."
            })
            continue

        # Validación de fecha límite
        if fecha_rec > fecha_limite:
            filas_omitidas_fecha += 1
            errores.append({
                "fila": num_fila,
                "dni": dni_str,
                "error": f"La fecha {fecha_rec} supera la fecha límite configurada ({fecha_limite}). Omitida."
            })
            continue

        try:
            hora_ent = _parse_time(entrada_val)
            dt_entrada = datetime.combine(fecha_rec, hora_ent)
        except ValueError as ve:
            errores.append({"fila": num_fila, "dni": dni_str, "error": str(ve)})
            continue

        dt_salida = None
        if salida_val is not None and str(salida_val).strip() != "":
            try:
                hora_sal = _parse_time(salida_val)
                dt_salida = datetime.combine(fecha_rec, hora_sal)
            except ValueError as ve:
                errores.append({"fila": num_fila, "dni": dni_str, "error": str(ve)})
                continue

        # Buscar asistencia existente
        asistencia = db.query(Asistencia).filter(
            Asistencia.empleado_id == empleado.id,
            Asistencia.fecha == fecha_rec
        ).first()

        if asistencia:
            asistencia.hora_entrada = dt_entrada
            if dt_salida:
                asistencia.hora_salida = dt_salida
            asistencia.origen_entrada = "migracion"
            asistencia.origen_salida = "migracion" if dt_salida else asistencia.origen_salida
            asistencia.motivo = "Migración histórica de asistencias"
            asistencia.agente_id = "MIGRACION_ADMIN"
            filas_actualizadas += 1
        else:
            asistencia = Asistencia(
                empleado_id=empleado.id,
                fecha=fecha_rec,
                hora_entrada=dt_entrada,
                hora_salida=dt_salida,
                origen_entrada="migracion",
                origen_salida="migracion" if dt_salida else None,
                motivo="Migración histórica de asistencias",
                agente_id="MIGRACION_ADMIN"
            )
            db.add(asistencia)
            filas_creadas += 1

        filas_procesadas += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"ok": False, "error": f"Error al guardar asistencias en la base de datos: {str(e)}"}

    return {
        "ok": True,
        "filas_procesadas": filas_procesadas,
        "filas_creadas": filas_creadas,
        "filas_actualizadas": filas_actualizadas,
        "filas_omitidas_fecha": filas_omitidas_fecha,
        "total_errores": len(errores),
        "errores": errores
    }


def procesar_migracion_excel(db: Session, file_bytes: bytes, fecha_limite: date, filename: str = "migracion.xlsx") -> Dict[str, Any]:
    """
    Wrapper retrocompatible para procesar_migracion_archivo.
    """
    return procesar_migracion_archivo(db, file_bytes, filename, fecha_limite)

